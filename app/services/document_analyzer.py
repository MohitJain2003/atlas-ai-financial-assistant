"""
Document & Voice Analyzers - Document intelligence (PDF/DOCX) and Voice processing.
"""
import os
import logging
from typing import Dict, Any
from PyPDF2 import PdfReader
import docx

from app.ai.models import model_chain

logger = logging.getLogger(__name__)


class DocumentAnalyzerService:
    """Service to extract text and analyze financial documents."""

    @staticmethod
    async def extract_text(file_path: str, file_type: str) -> str:
        """Extract plain text from uploaded PDF or Word documents."""
        try:
            if file_type == "pdf":
                reader = PdfReader(file_path)
                text = ""
                for page in reader.pages:
                    extracted = page.extract_text()
                    if extracted:
                        text += extracted + "\n"
                return text.strip()

            elif file_type in ("docx", "doc"):
                doc = docx.Document(file_path)
                return "\n".join([p.text for p in doc.paragraphs if p.text]).strip()

            elif file_type in ("xlsx", "xls"):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                lines = []
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    lines.append(f"--- Sheet: {sheetname} ---")
                    for row in sheet.iter_rows(values_only=True):
                        if any(row):
                            row_vals = [str(cell) if cell is not None else "" for cell in row]
                            lines.append(" | ".join(row_vals))
                return "\n".join(lines).strip()

            elif file_type == "csv":
                import csv
                lines = []
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if any(row):
                            lines.append(" | ".join(row))
                return "\n".join(lines).strip()

            elif file_type == "txt":
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    return f.read().strip()

        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")

        return ""

    @staticmethod
    async def analyze_document(user_prompt: str, document_text: str) -> str:
        """Analyze financial document text using AI."""
        if not document_text:
            return "Could not extract readable text from the uploaded document."

        # Truncate text if too long (max ~15,000 chars for prompt safety)
        truncated_text = document_text[:15000]
        if len(document_text) > 15000:
            truncated_text += "\n\n[Document truncated for length...]"

        system_instruction = (
            "You are Atlas, an expert financial analyst. Analyze the following document snippet "
            "and answer the user's questions concisely. Focus on key financial KPIs, risks, "
            "revenue trends, and actionable takeaways."
        )

        user_message = f"User Request: {user_prompt}\n\nDocument Text:\n{truncated_text}"

        return await model_chain.generate(system_instruction, user_message)


class VoiceService:
    """Service to handle voice message processing."""

    @staticmethod
    async def process_voice_message(file_path: str) -> str:
        """Transcribe voice message audio to text."""
        transcription = await model_chain.transcribe_voice(file_path)
        if not transcription:
            return "Sorry, I couldn't process or transcribe your voice message."
        return transcription
